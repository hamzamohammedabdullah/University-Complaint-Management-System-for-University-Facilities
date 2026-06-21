from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.forms import PasswordChangeForm
from django.contrib.auth import update_session_auth_hash
from .forms import FacilityUserRegistrationForm, LoginForm, CreateStaffForm, AltLoginForm
from .models import User
from django.conf import settings
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

@login_required
def register_view(request):
    # Allow Maintenance Staff and admins to create new facility user accounts.
    if not (request.user.is_staff_member or request.user.is_admin):
        messages.error(request, 'Only Maintenance Staff and admins can create user accounts.')
        return redirect('dashboard')

    form = FacilityUserRegistrationForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.save()
        # send default password to user's email
        try:
            from complaints.notifications import send_email_notification
            default_pw = getattr(settings, 'DEFAULT_FACILITY_PASSWORD', 'ChangeMe123!')
            subject = 'Your new account has been created'
            msg = f"Your account ({user.username}) has been created. Your temporary password is: {default_pw}\nPlease change it after signing in."
            send_email_notification(getattr(user, 'email', None), subject, msg)
        except Exception:
            pass
        # Do not auto-login the newly created user; staff created the account.
        messages.success(request, f'User {user.first_name or user.username} has been created.')
        return redirect('admin_users')
    return render(request, 'accounts/register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = LoginForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        # Detect first login by checking previous last_login
        first_login = user.last_login is None
        login(request, user)
        messages.success(request, f'Welcome back, {user.first_name or user.username}!')

        if first_login:
            try:
                from complaints.models import Notification
                from complaints.notifications import send_email_notification
                title = 'Please update your password'
                message = 'It looks like this is your first time signing in. For security, please update your password from your profile settings.'
                Notification.objects.create(recipient=user, complaint=None, title=title, message=message, notif_type='general')
                send_email_notification(getattr(user, 'email', None), title, message)
            except Exception:
                pass

        return redirect(request.GET.get('next', 'dashboard'))
    return render(request, 'accounts/login.html', {'form': form})


def alt_login_view(request):
    """Alternative login page with a role dropdown.

    Does not replace the existing login. If the selected role does not
    match the authenticated user's role, the login is rejected with
    a form error.
    """
    if request.user.is_authenticated:
        return redirect('dashboard')
    form = AltLoginForm(request, data=request.POST or None)
    if request.method == 'POST' and form.is_valid():
        user = form.get_user()
        selected_role = form.cleaned_data.get('role')
        if selected_role and user.role != selected_role:
            form.add_error(None, 'Selected role does not match this account.')
        else:
            login(request, user)
            messages.success(request, f'Welcome back, {user.first_name or user.username}!')
            return redirect(request.GET.get('next', 'dashboard'))
    return render(request, 'accounts/login_alt.html', {'form': form})

def logout_view(request):
    logout(request)
    messages.info(request, 'You have been signed out.')
    return redirect('login')

@login_required
def profile_view(request):
    if request.method == 'POST':
        user = request.user
        user.first_name = request.POST.get('first_name', user.first_name)
        user.last_name  = request.POST.get('last_name', user.last_name)
        user.phone      = request.POST.get('phone', user.phone)
        user.save()
        messages.success(request, 'Profile updated successfully.')
        return redirect('profile')
    return render(request, 'accounts/profile.html')


@login_required
def change_password_view(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)  # keep the user logged in
            messages.success(request, 'Your password has been updated.')
            return redirect('profile')
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'accounts/change_password.html', {'form': form})

# Admin: list users
@login_required
def admin_users_view(request):
    # Allow Maintenance Staff and admins to access this page
    if not (request.user.is_admin or request.user.is_staff_member):
        messages.error(request, 'Access denied.')
        return redirect('dashboard')

    role_filter = request.GET.get('role', '')
    users = User.objects.all().order_by('-date_joined')
    if role_filter:
        users = users.filter(role=role_filter)

    # Inline student creation form for maintenance staff and admins
    create_form = None
    if request.user.is_staff_member or request.user.is_admin:
        from .forms import FacilityUserRegistrationForm
        if request.method == 'POST' and request.POST.get('create_student'):
            create_form = FacilityUserRegistrationForm(request.POST)
            if create_form.is_valid():
                new_user = create_form.save()
                # send default password to new facility user
                try:
                    from complaints.notifications import send_email_notification
                    default_pw = getattr(settings, 'DEFAULT_FACILITY_PASSWORD', 'ChangeMe123!')
                    subject = 'Your new account has been created'
                    msg = f"Your account ({new_user.username}) has been created. Your temporary password is: {default_pw}\nPlease change it after signing in."
                    send_email_notification(getattr(new_user, 'email', None), subject, msg)
                except Exception:
                    pass
                messages.success(request, f'Student account {new_user.get_full_name() or new_user.username} created.')
                return redirect('admin_users')
        else:
            create_form = FacilityUserRegistrationForm()
        # Handle batch Excel upload for creating multiple facility users
        if request.method == 'POST' and request.FILES.get('batch_file'):
            batch_file = request.FILES.get('batch_file')
            if load_workbook is None:
                messages.error(request, 'Excel import requires openpyxl. Please install the dependency.')
            else:
                try:
                    wb = load_workbook(filename=batch_file, data_only=True)
                    ws = wb.active
                    header = [str(c).strip().lower() if c is not None else '' for c in next(ws.iter_rows(values_only=True))]
                    # mapping expected headers
                    expected = ['first_name', 'last_name', 'username', 'email', 'facility_id', 'phone']
                    idx = {h: header.index(h) for h in header if h in expected}
                    created = []
                    skipped = []
                    default_pw = getattr(settings, 'DEFAULT_FACILITY_PASSWORD', 'ChangeMe123!')
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        username = row[idx.get('username')] if idx.get('username') is not None else None
                        email = row[idx.get('email')] if idx.get('email') is not None else None
                        if not username or not email:
                            skipped.append((username or '', 'missing username/email'))
                            continue
                        if User.objects.filter(username=username).exists():
                            skipped.append((username, 'already exists'))
                            continue
                        user = User(username=username, email=email)
                        user.first_name = row[idx.get('first_name')] or '' if idx.get('first_name') is not None else ''
                        user.last_name = row[idx.get('last_name')] or '' if idx.get('last_name') is not None else ''
                        user.facility_id = row[idx.get('facility_id')] or '' if idx.get('facility_id') is not None else ''
                        user.phone = row[idx.get('phone')] or '' if idx.get('phone') is not None else ''
                        user.role = 'facility'
                        user.set_password(default_pw)
                        user.save()
                        try:
                            from complaints.notifications import send_email_notification
                            subject = 'Your new account has been created'
                            msg = f"Your account ({user.username}) has been created. Your temporary password is: {default_pw}\nPlease change it after signing in."
                            send_email_notification(getattr(user, 'email', None), subject, msg)
                        except Exception:
                            pass
                        created.append(user.username)
                    messages.success(request, f'Imported {len(created)} users; skipped {len(skipped)} rows.')
                except Exception as e:
                    messages.error(request, f'Failed to import Excel file: {e}')

    return render(request, 'accounts/admin_users.html', {'users': users, 'role_filter': role_filter, 'create_form': create_form})

# Admin: create staff/admin account
@login_required
def create_user_view(request):
    if not request.user.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    form = CreateStaffForm(request.POST or None)
    # Remove 'facility' choice for this admin/staff creation form
    form.fields['role'].choices = [c for c in form.fields['role'].choices if c[0] != 'facility']
    # Prevent ordinary admins from creating other admin accounts; only superusers may create admins
    if not request.user.is_superuser:
        # Filter out 'admin' choice from the role field
        form.fields['role'].choices = [c for c in form.fields['role'].choices if c[0] != 'admin']

    if request.method == 'POST' and form.is_valid():
        selected_role = form.cleaned_data.get('role')
        if selected_role == 'admin' and not request.user.is_superuser:
            form.add_error('role', 'Only superusers may create administrator accounts.')
        else:
            new_user = form.save()
            # send default password email
            try:
                from complaints.notifications import send_email_notification
                if new_user.role == 'facility':
                    default_pw = getattr(settings, 'DEFAULT_FACILITY_PASSWORD', 'ChangeMe123!')
                else:
                    default_pw = getattr(settings, 'DEFAULT_STAFF_ADMIN_PASSWORD', 'AdminChangeMe123!')
                subject = 'Your new account has been created'
                msg = f"Your account ({new_user.username}) has been created. Your temporary password is: {default_pw}\nPlease change it after signing in."
                send_email_notification(getattr(new_user, 'email', None), subject, msg)
            except Exception:
                pass
            messages.success(request, 'User account created successfully.')
            return redirect('admin_users')

    return render(request, 'accounts/create_user.html', {'form': form})

# Admin: toggle active
@login_required
def toggle_user_view(request, user_id):
    if not request.user.is_admin:
        messages.error(request, 'Access denied.')
        return redirect('dashboard')
    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()
    status = 'activated' if user.is_active else 'deactivated'
    messages.success(request, f'User {user.get_full_name()} has been {status}.')
    return redirect('admin_users')
